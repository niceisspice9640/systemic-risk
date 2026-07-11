from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "data/cloud_hhi.xlsx"

# ── Palette ────────────────────────────────────────────────────────────────────
C_HDR_BG  = "1F3864"   # dark navy header
C_HDR2_BG = "344D6E"   # section sub-header
C_HDR_FG  = "FFFFFF"
C_INPUT   = "0000FF"   # hardcoded inputs (industry standard)
C_FORMULA = "000000"   # formulas
C_XSHEET  = "008000"   # cross-sheet references
C_NOTE_BG = "FFFBCC"   # yellow — notes
C_STRIPE  = "F2F7FF"   # alternating row
C_FLAG_BG = "FFEAEA"   # red — 2024 flag
C_WHITE   = "FFFFFF"

def font(color=C_FORMULA, bold=False, size=9, italic=False, name="Arial"):
    return Font(name=name, color=color, bold=bold, size=size, italic=italic)

def fill(color):
    return PatternFill("solid", start_color=color, fgColor=color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_cell(ws, row, col, value, fnt, fll, num_fmt=None,
               halign="left", valign="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = fnt
    c.fill = fll
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    c.border = thin_border()
    if num_fmt:
        c.number_format = num_fmt
    return c

def hdr_row(ws, row, cols, values, bg=C_HDR_BG, h=16):
    for col, val in zip(cols, values):
        apply_cell(ws, row, col, val,
                   font(C_HDR_FG, bold=True, size=9),
                   fill(bg), halign="center")
    ws.row_dimensions[row].height = h

def section_banner(ws, row, col_start, col_end, text, bg=C_HDR2_BG):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    apply_cell(ws, row, col_start, text,
               font(C_HDR_FG, bold=True, size=9), fill(bg))
    ws.row_dimensions[row].height = 16

def title_row(ws, row, col_start, col_end, text, h=24):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    apply_cell(ws, row, col_start, text,
               font(C_HDR_FG, bold=True, size=12), fill(C_HDR_BG),
               halign="center")
    ws.row_dimensions[row].height = h

# ── Workbook ───────────────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Data
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Data"

for i, w in enumerate([8, 11, 11, 11, 11, 11, 58], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title
title_row(ws1, 1, 1, 7,
          "Variable 2 — Cloud Infrastructure Concentration: HHI Analysis (2017–2024)")

# Subtitle
ws1.merge_cells("A2:G2")
apply_cell(ws1, 2, 1,
           "Research: \"To what extent has the growth of cyberattacks and infrastructure "
           "disruptions outpaced institutional defensive and regulatory capacity?\" (2015–2025)",
           font("555555", italic=True, size=9), fill("F5F5F5"),
           halign="left", wrap=True)
ws1.row_dimensions[2].height = 16

# Color key
ws1.merge_cells("A3:G3")
apply_cell(ws1, 3, 1,
           "Color key:  Blue = hardcoded input    Black = formula    Green = cross-sheet ref    "
           "Red background = methodological flag",
           font("666666", italic=True, size=8), fill("F5F5F5"), halign="left")
ws1.row_dimensions[3].height = 14

# Column headers (row 4)
col_headers = ["Year", "AWS (%)", "Azure (%)", "GCP (%)", "CR3 (%)", "HHI", "Source Note"]
hdr_row(ws1, 4, range(1, 8), col_headers, h=17)

# Data
rows = [
    (2017, 33, 12, 6,
     "Primary — Synergy Research Group; Canalys Q4 2017 (GCP baseline confirmed)"),
    (2018, 33, 13, 7,
     "Interpolated — Synergy Q1 2022 press release: Azure +~2 pp/yr, GCP +~1 pp/yr stated rate"),
    (2019, 33, 15, 8,
     "Interpolated — Synergy Q1 2022 press release rate-of-change"),
    (2020, 33, 17, 8,
     "Interpolated — Synergy Q1 2022 press release rate-of-change"),
    (2021, 33, 19, 9,
     "Interpolated — Synergy Q1 2022 press release rate-of-change"),
    (2022, 33, 21, 9,
     "Primary — Synergy Research Group; businesstats.com CR3 cross-check (63% confirmed)"),
    (2023, 31, 24, 11,
     "Primary — AAG-IT citing Synergy (AWS 31%, Azure 23%+); holori.com GCP cross-check"),
    (2024, 30, 21, 12,
     "Primary — Synergy Research Group Q4 2024 press release  †RECLASSIFICATION — see note below"),
]

for r_idx, (yr, aws, azure, gcp, note) in enumerate(rows, 5):
    is_flag   = (yr == 2024)
    is_stripe = (r_idx % 2 == 0) and not is_flag
    rf = fill(C_FLAG_BG) if is_flag else (fill(C_STRIPE) if is_stripe else fill(C_WHITE))

    apply_cell(ws1, r_idx, 1, yr,
               font(C_INPUT, bold=True), rf, "0", halign="center")
    apply_cell(ws1, r_idx, 2, aws,
               font(C_INPUT), rf, "0", halign="center")
    apply_cell(ws1, r_idx, 3, azure,
               font(C_INPUT), rf, "0", halign="center")
    apply_cell(ws1, r_idx, 4, gcp,
               font(C_INPUT), rf, "0", halign="center")
    apply_cell(ws1, r_idx, 5, f"=B{r_idx}+C{r_idx}+D{r_idx}",
               font(C_FORMULA), rf, "0", halign="center")
    apply_cell(ws1, r_idx, 6, f"=B{r_idx}^2+C{r_idx}^2+D{r_idx}^2",
               font(C_FORMULA), rf, "#,##0", halign="center")
    note_color = "993333" if is_flag else "555555"
    apply_cell(ws1, r_idx, 7, note,
               font(note_color, italic=is_flag, size=9), rf,
               halign="left", wrap=True)
    ws1.row_dimensions[r_idx].height = 15

# Notes section (rows 14–18)
ws1.row_dimensions[13].height = 8
section_banner(ws1, 14, 1, 7, "METHODOLOGY NOTES")

note_data = [
    ("†",
     "2024 RECLASSIFICATION: Synergy Research Group reclassified a portion of Azure revenues "
     "from IaaS/PaaS to SaaS in 2024, producing an apparent drop in Azure share from 24% (2023) "
     "to 21% (2024) despite continued absolute revenue growth. The 2024 HHI (1,485) reflects "
     "this artifact. Primary trend is 2017–2023 for OLS purposes."),
    ("HHI",
     "Formula: Σ(sᵢ²) where sᵢ = each provider's market share as whole-number percentage points. "
     "Big Three only. Full-market HHI follows the same directional trend."),
    ("CR3",
     "Concentration Ratio (top 3) = AWS + Azure + GCP combined share. Rising from 51% (2017) "
     "to 66% (2023) corroborates HHI trend independently."),
    ("Sources",
     "Synergy Research Group press releases: srgresearch.com/articles (free, no paywall). "
     "Canalys Q4 2017. holori.com/cloud-market-share-2024-aws-azure-gcp (Synergy-citing). "
     "2018–2021 interpolated from Synergy Q1 2022 stated annual rate-of-change."),
]

for n_idx, (label, text) in enumerate(note_data, 15):
    apply_cell(ws1, n_idx, 1, label,
               font("333333", bold=True, size=9), fill(C_NOTE_BG),
               halign="center", valign="top")
    ws1.merge_cells(
        start_row=n_idx, start_column=2,
        end_row=n_idx, end_column=7
    )
    apply_cell(ws1, n_idx, 2, text,
               font("444444", size=9), fill(C_NOTE_BG),
               halign="left", valign="top", wrap=True)
    ws1.row_dimensions[n_idx].height = 40

ws1.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — OLS Regression
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("OLS Regression")

for i, w in enumerate([32, 18, 46, 38], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

title_row(ws2, 1, 1, 4, "OLS Regression: HHI ~ Year (2017–2023)")

ws2.merge_cells("A2:D2")
apply_cell(ws2, 2, 1,
           "Dependent variable: HHI (Big Three cloud concentration index). "
           "Independent variable: Year. n=7. 2024 excluded — Azure reclassification artifact.",
           font("555555", italic=True, size=9), fill("F5F5F5"),
           halign="left", wrap=True)
ws2.row_dimensions[2].height = 16

ws2.merge_cells("A3:D3")
apply_cell(ws2, 3, 1,
           "Color key:  Black = formula    Green = cross-sheet reference",
           font("666666", italic=True, size=8), fill("F5F5F5"), halign="left")
ws2.row_dimensions[3].height = 14

section_banner(ws2, 4, 1, 4, "REGRESSION OUTPUT")

hdr_row(ws2, 5, range(1, 5),
        ["Statistic", "Value", "Excel Formula", "Interpretation"])

# Row index tracking: stats start at row 6
# B6=slope, B7=intercept, B8=R², B9=STEYX, B10=n, B11=DEVSQ, B12=SE_slope, B13=t, B14=p
reg_stats = [
    ("β₁ — Slope (HHI / year)",
     "=SLOPE(Data!F5:F11,Data!A5:A11)",
     "=SLOPE(Data!F5:F11,Data!A5:A11)",
     "Avg. annual HHI increase — the primary trend coefficient",
     "#,##0.00"),
    ("β₀ — Intercept",
     "=INTERCEPT(Data!F5:F11,Data!A5:A11)",
     "=INTERCEPT(Data!F5:F11,Data!A5:A11)",
     "Theoretical HHI at year 0 (no standalone interpretation)",
     "#,##0.00"),
    ("R² (Coefficient of Determination)",
     "=RSQ(Data!F5:F11,Data!A5:A11)",
     "=RSQ(Data!F5:F11,Data!A5:A11)",
     "Proportion of HHI variance explained by year alone",
     "0.0000"),
    ("SE Regression — STEYX",
     "=STEYX(Data!F5:F11,Data!A5:A11)",
     "=STEYX(Data!F5:F11,Data!A5:A11)",
     "Average size of residuals — goodness of fit",
     "#,##0.00"),
    ("n (observations used)",
     "=COUNT(Data!F5:F11)",
     "=COUNT(Data!F5:F11)",
     "2017–2023 only (2024 excluded)",
     "0"),
    ("DEVSQ(x) — Σ(yearᵢ − ȳear)²",
     "=DEVSQ(Data!A5:A11)",
     "=DEVSQ(Data!A5:A11)",
     "Sum of squared deviations of year values from mean year",
     "#,##0.00"),
    ("SE(β₁) — Standard Error of Slope",
     "=B9/SQRT(B11)",
     "B9/SQRT(B11)  [STEYX / SQRT(DEVSQ(x))]",
     "Sampling uncertainty around the slope estimate",
     "#,##0.0000"),
    ("t-statistic (β₁)",
     "=B6/B12",
     "B6/B12  [slope / SE(slope)]",
     "Tests whether slope differs significantly from zero",
     "#,##0.00"),
    ("p-value (two-tailed)",
     "=TDIST(ABS(B13),B10-2,2)",
     "TDIST(ABS(B13), B10-2, 2)  [two-tailed, df = n-2]",
     "Probability of this trend occurring by chance; <0.05 = significant",
     "0.0000"),
]

# Data rows on Data sheet: 2017=row5, 2018=row6,...,2023=row11, 2024=row12
# So OLS uses Data!A5:A11 (years) and Data!F5:F11 (HHI)

for r_offset, (stat, formula, disp, interp, fmt) in enumerate(reg_stats):
    r = 6 + r_offset
    is_stripe = r % 2 == 0
    rf = fill(C_STRIPE) if is_stripe else fill(C_WHITE)

    apply_cell(ws2, r, 1, stat, font("333333", bold=(r_offset in [0, 2, 8])), rf,
               halign="left")
    apply_cell(ws2, r, 2, formula, font(C_XSHEET), rf, fmt, halign="center")
    apply_cell(ws2, r, 3, disp,
               font("777777", size=8.5, name="Courier New"), rf, halign="left")
    apply_cell(ws2, r, 4, interp, font("555555", italic=True, size=9), rf,
               halign="left")
    ws2.row_dimensions[r].height = 16

# Predicted vs Actual (starts at row 17)
ws2.row_dimensions[15].height = 8
section_banner(ws2, 16, 1, 4, "PREDICTED vs ACTUAL HHI — 2017–2023")
hdr_row(ws2, 17, range(1, 5), ["Year", "Actual HHI", "Predicted HHI", "Residual"])

# Data rows on Data sheet: 2017=row5 through 2023=row11
for p_offset, data_row in enumerate(range(5, 12)):
    pr = 18 + p_offset
    is_stripe = pr % 2 == 0
    rf = fill(C_STRIPE) if is_stripe else fill(C_WHITE)

    apply_cell(ws2, pr, 1, f"=Data!A{data_row}",
               font(C_XSHEET), rf, "0", halign="center")
    apply_cell(ws2, pr, 2, f"=Data!F{data_row}",
               font(C_XSHEET), rf, "#,##0", halign="center")
    # predicted = β₀ + β₁ × year  →  $B$7 + $B$6 * year
    apply_cell(ws2, pr, 3, f"=$B$7+$B$6*Data!A{data_row}",
               font(C_FORMULA), rf, "#,##0.00", halign="center")
    apply_cell(ws2, pr, 4, f"=B{pr}-C{pr}",
               font(C_FORMULA), rf, "#,##0.00", halign="center")
    ws2.row_dimensions[pr].height = 15

# Interpretation notes
ws2.row_dimensions[25].height = 8
section_banner(ws2, 26, 1, 4, "INTERPRETATION NOTES")

interp_notes = [
    "β₁: For every one-year increase in the observation window, the Big Three cloud HHI "
    "increased by approximately [β₁] points on average. Report as: "
    "\"HHI increased by X points/year (β₁=X, R²=X, p<0.05).\"",

    "R²: A value close to 1.0 indicates year alone explains most of the variance in HHI — "
    "the concentration trend is strongly linear over 2017–2023. Expected >0.95.",

    "p-value: Expected <0.001 despite n=7, because the trend is very consistent. "
    "Acknowledged limitation: low statistical power. Treat as descriptive-quantitative "
    "rather than strict inferential claim.",

    "Limitation: n=7 data points. OLS is underpowered for strong inference. Report β₁ "
    "with SE and p-value, and note limited statistical power explicitly in the paper's "
    "methodology section.",
]

for n_idx, note in enumerate(interp_notes, 27):
    ws2.merge_cells(
        start_row=n_idx, start_column=1,
        end_row=n_idx, end_column=4
    )
    apply_cell(ws2, n_idx, 1, note,
               font("444444", italic=True, size=9), fill(C_NOTE_BG),
               halign="left", valign="top", wrap=True)
    ws2.row_dimensions[n_idx].height = 38

# ── Non-parametric trend tests (rows 32+) ──────────────────────────────────────
ws2.row_dimensions[31].height = 8
section_banner(ws2, 32, 1, 4,
               "NON-PARAMETRIC TREND TESTS (small-n appropriate; see trend_tests.py)")
hdr_row(ws2, 33, range(1, 5),
        ["Statistic", "Value", "Method", "Interpretation"])

# Values computed in trend_tests.py (Mann-Kendall + Sen's slope), documented here
# as reproducible hardcoded results (blue), sourced to the script.
np_rows = [
    ("Mann-Kendall S", 21, "0",
     "sum of pairwise sign comparisons",
     "Signed count; positive = upward trend"),
    ("Kendall's tau", 1.0000, "0.0000",
     "S / [n(n-1)/2]",
     "1.00 = perfectly monotonic (every year > prior year)"),
    ("Mann-Kendall Z", 3.0038, "0.0000",
     "continuity-corrected standardized stat",
     "Test statistic for significance"),
    ("Mann-Kendall p-value", 0.0027, "0.0000",
     "two-tailed, normal approximation",
     "Significant at 0.01 level despite n=7"),
    ("Sen's slope", 70.00, "#,##0.00",
     "median of all pairwise slopes",
     "Non-parametric trend estimate: +70 HHI/year"),
    ("OLS slope (comparison)", 68.86, "#,##0.00",
     "least-squares slope (cell B6 above)",
     "Agreement with Sen's slope validates the estimate"),
]

for r_offset, (stat, val, fmt, method, interp) in enumerate(np_rows):
    r = 34 + r_offset
    is_stripe = r % 2 == 0
    rf = fill(C_STRIPE) if is_stripe else fill(C_WHITE)
    is_key = stat in ("Kendall's tau", "Mann-Kendall p-value", "Sen's slope")

    apply_cell(ws2, r, 1, stat, font("333333", bold=is_key), rf, halign="left")
    apply_cell(ws2, r, 2, val, font(C_INPUT, bold=is_key), rf, fmt, halign="center")
    apply_cell(ws2, r, 3, method,
               font("777777", size=8.5, name="Courier New"), rf, halign="left")
    apply_cell(ws2, r, 4, interp, font("555555", italic=True, size=9), rf,
               halign="left")
    ws2.row_dimensions[r].height = 16

ws2.merge_cells(start_row=40, start_column=1, end_row=40, end_column=4)
apply_cell(ws2, 40, 1,
           "Why non-parametric: with n=7, OLS significance rests on assumptions "
           "(normality, independence) that a short autocorrelated series may violate. "
           "The Mann-Kendall test makes no distributional assumption and is the standard "
           "trend test for small annual series; Sen's slope is its matched slope estimator. "
           "Values are hardcoded (blue) from the reproducible script trend_tests.py. Because "
           "tau=1.00 (perfect monotonicity), the trend clears significance despite small n. "
           "Caveat: a non-significant Mann-Kendall result on n=7 would indicate an "
           "under-powered series, not the absence of a trend.",
           font("444444", size=9), fill(C_NOTE_BG),
           halign="left", valign="top", wrap=True)
ws2.row_dimensions[40].height = 74

ws2.freeze_panes = "A6"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Key Findings
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Key Findings")

for i, w in enumerate([30, 18, 18, 38], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

title_row(ws3, 1, 1, 4, "Variable 2 — Key Findings Summary")

ws3.merge_cells("A2:D2")
apply_cell(ws3, 2, 1,
           "Cloud Infrastructure Market Concentration (HHI Analysis) | "
           "Source: Synergy Research Group quarterly press releases (srgresearch.com/articles)",
           font("555555", italic=True, size=9), fill("F5F5F5"),
           halign="left", wrap=True)
ws3.row_dimensions[2].height = 16

ws3.row_dimensions[3].height = 8
section_banner(ws3, 4, 1, 4, "HEADLINE METRICS (2017 vs 2023)")
hdr_row(ws3, 5, range(1, 5),
        ["Metric", "2017 (baseline)", "2023 (peak)", "Change"])

# Data rows reference Data sheet rows: 2017=row5, 2023=row11
kf_rows = [
    ("HHI (Big Three)",
     "=Data!F5", "=Data!F11", "=(C6-B6)/B6",
     "#,##0", "#,##0", "0.0%"),
    ("CR3 Combined Share",
     "=Data!E5", "=Data!E11", "=C7-B7",
     '0" pp"', '0" pp"', '0" pp"'),
    ("AWS Market Share",
     "=Data!B5", "=Data!B11", "=C8-B8",
     '0" pp"', '0" pp"', '0" pp"'),
    ("Azure Market Share",
     "=Data!C5", "=Data!C11", "=C9-B9",
     '0" pp"', '0" pp"', '0" pp"'),
    ("GCP Market Share",
     "=Data!D5", "=Data!D11", "=C10-B10",
     '0" pp"', '0" pp"', '0" pp"'),
]

for r_offset, (metric, base_f, peak_f, chg_f, fmt_b, fmt_p, fmt_c) in enumerate(kf_rows):
    r = 6 + r_offset
    is_stripe = r % 2 == 0
    rf = fill(C_STRIPE) if is_stripe else fill(C_WHITE)
    is_bold = (r_offset == 0)

    apply_cell(ws3, r, 1, metric,
               font("222222", bold=is_bold), rf, halign="left")
    apply_cell(ws3, r, 2, base_f,
               font(C_XSHEET), rf, fmt_b, halign="center")
    apply_cell(ws3, r, 3, peak_f,
               font(C_XSHEET), rf, fmt_p, halign="center")
    apply_cell(ws3, r, 4, chg_f,
               font(C_FORMULA, bold=is_bold), rf, fmt_c, halign="center")
    ws3.row_dimensions[r].height = 16

ws3.row_dimensions[11].height = 8
section_banner(ws3, 12, 1, 4, "TREND ANALYSIS SUMMARY (see 'OLS Regression' sheet)")
hdr_row(ws3, 13, range(1, 5),
        ["Statistic", "Value", "", "Note"])

ols_rows = [
    ("β₁ — OLS Slope (HHI/year)",
     "='OLS Regression'!B6", "#,##0.00",
     "Avg. annual HHI increase, 2017–2023"),
    ("R² (variance explained)",
     "='OLS Regression'!B8", "0.0000",
     "Proportion of variance explained by year (OLS)"),
    ("Sen's slope (HHI/year)",
     "='OLS Regression'!B38", "#,##0.00",
     "Non-parametric slope; agrees with OLS, validating the estimate"),
    ("Kendall's tau",
     "='OLS Regression'!B35", "0.0000",
     "1.00 = perfectly monotonic increase, 2017–2023"),
    ("Mann-Kendall p-value",
     "='OLS Regression'!B37", "0.0000",
     "Non-parametric trend test; significant at 0.01 despite n=7"),
    ("n (observations)",
     "='OLS Regression'!B10", "0",
     "2017–2023 only; 2024 excluded due to reclassification"),
]

for r_offset, (stat, val_f, fmt, note) in enumerate(ols_rows):
    r = 14 + r_offset
    is_stripe = r % 2 == 0
    rf = fill(C_STRIPE) if is_stripe else fill(C_WHITE)

    apply_cell(ws3, r, 1, stat, font("333333"), rf, halign="left")
    apply_cell(ws3, r, 2, val_f, font(C_XSHEET), rf, fmt, halign="center")
    apply_cell(ws3, r, 3, "", font("333333"), rf)
    apply_cell(ws3, r, 4, note, font("555555", italic=True, size=9), rf,
               halign="left")
    ws3.row_dimensions[r].height = 16

ws3.row_dimensions[20].height = 8
section_banner(ws3, 21, 1, 4, "RESEARCH CONTEXT")

context_notes = [
    ("Variable role",
     "Variable 2 of 3. Measures structural blast radius — not outage frequency "
     "directly, but the concentration of infrastructure that amplifies the impact "
     "of any single failure (malicious or non-malicious)."),
    ("Theoretical link",
     "Rising HHI operationalises the efficiency-fragility tradeoff: as digital "
     "workloads concentrate in fewer providers for efficiency, the blast radius "
     "of any cascade or outage grows proportionally."),
    ("Case study connection",
     "Connects to: 2022 Kakao data center fire (single-facility SPOF), "
     "2024 CrowdStrike update cascade (single-vendor failure → global aviation, "
     "finance, healthcare), Canvas/Instructure LMS (platform-dependency cascade)."),
    ("Limitations",
     "n=7 for OLS/Mann-Kendall (low statistical power). 2024 HHI affected by Synergy's "
     "Azure reclassification. 2018–2021 shares interpolated, not confirmed from primary "
     "Synergy press releases. All limitations noted explicitly in paper methodology."),
]

for n_idx, (label, text) in enumerate(context_notes, 22):
    apply_cell(ws3, n_idx, 1, label,
               font("333333", bold=True, size=9), fill(C_NOTE_BG),
               halign="left", valign="top")
    ws3.merge_cells(
        start_row=n_idx, start_column=2,
        end_row=n_idx, end_column=4
    )
    apply_cell(ws3, n_idx, 2, text,
               font("444444", size=9), fill(C_NOTE_BG),
               halign="left", valign="top", wrap=True)
    ws3.row_dimensions[n_idx].height = 44

wb.save(OUTPUT)
print(f"✓ Saved: {OUTPUT}")
